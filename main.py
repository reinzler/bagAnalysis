from pathlib import Path
from rosbags.rosbag2 import Reader
from rosbags.serde import deserialize_cdr
from rosbags.image import message_to_cvimage
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import math
import os
import numpy as np
from PIL import Image, ImageDraw, ImageFont
from tqdm import tqdm
from openpyxl import Workbook
import cv2


def extractData(output_directory):
    with Reader(Path(input_directory)) as reader:
        fix_longitude, fix_latitude, fix_altitude, time_epoch, h_acc, v_acc = [], [], [], [], [], []
        fil_longitude, fil_latitude, fil_altitude, xmag, ymag, zmag = [], [], [], [], [], []
        w_pixhawk, x_pixhawk, y_pixhawk, z_pixhawk, pressure = [], [], [], [], []
        w_kalman, x_kalman, y_kalman, z_kalman = [], [], [], []
        angular_velocity_x, angular_velocity_y, angular_velocity_z = [], [], []
        linear_acceleration_x, linear_acceleration_y, linear_acceleration_z = [], [], []
        particle_longitude, particle_latitude, particle_altitude, particle_yaw = [], [], [], []
        particle_lat_rms, particle_long_rms, particle_alt_rms, particle_yaw_rms = [], [], [], []
        kalman_lat_rms, kalman_long_rms, kalman_alt_rms = [], [], []
        vx_kalman, vy_kalman, vz_kalman, wx_kalman, wy_kalman, wz_kalman = [], [], [], [], [], []
        vx_kalman_rms, vy_kalman_rms, vz_kalman_rms, wx_kalman_rms, wy_kalman_rms, wz_kalman_rms = [], [], [], [], [], []
        vx_pixhawk, vy_pixhawk, vz_pixhawk = [], [], []

        for connection, timestamp, rawdata in reader.messages():
            ### extracting GNSS raw data from /gps/fix topic
            if connection.topic == '/gps/fix':
                msg = deserialize_cdr(rawdata, connection.msgtype)
                fix_longitude.append(msg.longitude)
                fix_latitude.append(msg.latitude)
                fix_altitude.append(msg.altitude)
                timestamp = msg.header.stamp.sec + msg.header.stamp.nanosec / (10 ** len(str(msg.header.stamp.nanosec)))
                time_epoch.append(timestamp)

            ### extracting images from /RGB topic and conversion images to shades of gray /camera/color/image_raw
            if connection.topic == '/camera/color/image_raw':
                msg = deserialize_cdr(rawdata, connection.msgtype)
                timestamp = msg.header.stamp.sec + msg.header.stamp.nanosec / (10 ** len(str(msg.header.stamp.nanosec)))
                img = message_to_cvimage(msg)
                if os.path.exists(f"{output_directory}" + '/images'):
                    pass
                else:
                    os.mkdir(f"{output_directory}" + '/images')
                img_gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)  # Convert to gray
                output = f"{output_directory}" + f"/images/{timestamp}.png"
                cv2.imwrite(output, img_gray)

            ### extracting GNSS accuracy data from /gps_acc topic, where h_acc - horizontal
            # (position uncertainty (standard deviation)) accuracy in mm,
            # v_acc - vertical (position uncertainty (standard deviation)) accuracy in mm
            if connection.topic == '/gps_acc':
                msg = deserialize_cdr(rawdata, connection.msgtype)
                h_acc.append(msg.data[0])
                v_acc.append(msg.data[1])

            ### extracting velocities from /gbl/pos/int topic, MAVLink message - GLOBAL_POSITION_INT
            if connection.topic == '/gbl/pos/int':
                msg = deserialize_cdr(rawdata, connection.msgtype)
                vx_pixhawk.append(msg.x)
                vy_pixhawk.append(msg.y)
                vz_pixhawk.append(msg.z)

            ### extracting magnetometer data xmag, ymag, zmag from /Mag topic pixhawk/odometry.pose.pose.position.z
            if connection.topic == '/pixhawk/odometry':
                msg = deserialize_cdr(rawdata, connection.msgtype)
                xmag.append(msg.pose.pose.position.x)
                ymag.append(msg.pose.pose.position.y)
                zmag.append(msg.pose.pose.position.z)

            ### extracting orientation data (quaternion) from /RS_IMU topic
            if connection.topic == '/pixhawk/imu':  # RS_IMU, Pixhawk_IMU
                msg = deserialize_cdr(rawdata, connection.msgtype)
                w_pixhawk.append(msg.orientation.w)
                x_pixhawk.append(msg.orientation.x)
                y_pixhawk.append(msg.orientation.y)
                z_pixhawk.append(msg.orientation.z)
                angular_velocity_x.append(msg.angular_velocity.x)
                angular_velocity_y.append(msg.angular_velocity.y)
                angular_velocity_z.append(msg.angular_velocity.z)
                linear_acceleration_x.append(msg.linear_acceleration.x)
                linear_acceleration_y.append(msg.linear_acceleration.y)
                linear_acceleration_z.append(msg.linear_acceleration.z)

            ### extracting orientation (quaternion), linear and angular velocities from /odometry/filtered, after EKF!
            if connection.topic == '/odometry/filtered':  # RS_IMU, Pixhawk_IMU
                msg = deserialize_cdr(rawdata, connection.msgtype)
                x_kalman.append(msg.pose.pose.orientation.x)
                y_kalman.append(msg.pose.pose.orientation.y)
                z_kalman.append(msg.pose.pose.orientation.z)
                w_kalman.append(msg.pose.pose.orientation.w)
                vx_kalman.append(msg.twist.twist.linear.x)
                vy_kalman.append(msg.twist.twist.linear.y)
                vz_kalman.append(msg.twist.twist.linear.z)
                wx_kalman.append(msg.twist.twist.angular.x)
                wy_kalman.append(msg.twist.twist.angular.y)
                wz_kalman.append(msg.twist.twist.angular.z)
                vx_kalman_rms.append(math.sqrt(msg.twist.covariance[0]))
                vy_kalman_rms.append(math.sqrt(msg.twist.covariance[7]))
                vz_kalman_rms.append(math.sqrt(msg.twist.covariance[14]))
                wx_kalman_rms.append(math.sqrt(msg.twist.covariance[21]))
                wy_kalman_rms.append(math.sqrt(msg.twist.covariance[28]))
                wz_kalman_rms.append(math.sqrt(msg.twist.covariance[35]))

            ### extracting barometer data (pressure)  in hPa from /Baro topic
            if connection.topic == '/Baro':
                msg = deserialize_cdr(rawdata, connection.msgtype)
                pressure.append(msg.fluid_pressure)

            ### extracting GNSS filtered data after the Kalman filter fusion from /gps/filtered topic
            if connection.topic == '/gps/filtered':
                msg = deserialize_cdr(rawdata, connection.msgtype)
                fil_longitude.append(msg.longitude)
                fil_latitude.append(msg.latitude)
                fil_altitude.append(msg.altitude)
                kalman_lat_rms.append(math.sqrt(msg.position_covariance[0]))
                kalman_long_rms.append(math.sqrt(msg.position_covariance[4]))
                kalman_alt_rms.append(math.sqrt(msg.position_covariance[8]))

                ### extracting GNSS filtered data after the Particle filter from /particle_filter/mean_lla topic
            if connection.topic == '/particle_filter/mean_lla':
                msg = deserialize_cdr(rawdata, connection.msgtype)
                particle_longitude.append(msg.y)
                particle_latitude.append(msg.x)
                particle_altitude.append(msg.z)
                particle_yaw.append(msg.w)

            ### extracting RMS (std) for LLA and Yaw after the Particle filter from particle_filter/std_lla topic
            if connection.topic == '/particle_filter/std_lla':
                msg = deserialize_cdr(rawdata, connection.msgtype)
                particle_lat_rms.append(msg.x * 111000)
                particle_long_rms.append(msg.y * 111000)
                particle_alt_rms.append(msg.z)
                particle_yaw_rms.append(msg.w)

    return fix_longitude, fix_latitude, fix_altitude, time_epoch, fil_longitude, fil_latitude, fil_altitude, h_acc, \
        v_acc, xmag, ymag, zmag, w_pixhawk, x_pixhawk, y_pixhawk, z_pixhawk, w_kalman, x_kalman, y_kalman, z_kalman, \
        angular_velocity_x, angular_velocity_y, angular_velocity_z, linear_acceleration_x, linear_acceleration_y, \
        linear_acceleration_z, pressure, particle_longitude, particle_latitude, particle_altitude, particle_yaw, \
        particle_lat_rms, particle_long_rms, particle_alt_rms, particle_yaw_rms, kalman_lat_rms, kalman_long_rms, kalman_alt_rms, \
        vx_kalman, vy_kalman, vz_kalman, wx_kalman, wy_kalman, wz_kalman, vx_kalman_rms, vy_kalman_rms, vz_kalman_rms, wx_kalman_rms, \
        wy_kalman_rms, wz_kalman_rms, vx_pixhawk, vy_pixhawk, vz_pixhawk


def calculations(fix_longitude, fix_latitude, fix_altitude, time_epoch, fil_longitude, fil_latitude, fil_altitude,
                 x_pixhawk, y_pixhawk, z_pixhawk, w_pixhawk,
                 x_kalman, y_kalman, z_kalman, w_kalman, particle_yaw, particle_yaw_rms):
    """
        Function to calculate additional values to measure the performance of the navigation system.
        To find a difference (delta) between GNSS coordinates of a refeference and filtered results after EKF, in all
         of 3 axis - latitude, longitude, altitude and convert this difference from degrees to meters, we are using
          this formula:
            delta_coord = abs(fix_coord[i] - fil_coord[i]) * 111000
        To calculate horizontal:
            horizontal = math.sqrt((fil_longitude[i] - fix_longitude[i])**2 + (fil_latitude[i] - fix_latitude[i])**2))
        To calculate circle:
            circle = math.sqrt((fil_longitude[i] - fix_longitude[i]) ** 2 + (fil_latitude[i] - fix_latitude[i]) ** 2 +
                          (fil_altitude[i] - fix_altitude[i]) ** 2))
        To calculate SD (standard deviation):
            sd += math.sqrt((fil_longitude[i] - fix_longitude[i]) ** 2 + (fil_latitude[i] - fix_latitude[i]) ** 2
             + (fil_altitude[i] - fix_altitude[i]) ** 2)
        Also in this function where are a conversion of orientation from quaternion to angles in radians
        and degress
    """
    horizontal, circle, delta_time = [], [], []
    kalman_delta_longitude, kalman_delta_latitude, kalman_delta_altitude = [], [], []
    particle_delta_longitude, particle_delta_latitude, particle_delta_altitude = [], [], []
    delta_lon, delta_lat, std_kalman, std_particle = 0, 0, 0, 0
    pixhawk_quaternion = [x_pixhawk, y_pixhawk, z_pixhawk, w_pixhawk]
    kalman_quaternion = [x_kalman, y_kalman, z_kalman, w_kalman]
    delta_vx, delta_vy, delta_vz = [], [], []

    print(len(fix_longitude), len(fil_longitude))
    for i in range(len(fix_longitude)):
        delta_lon = abs(fix_longitude[i] - fil_longitude[i]) * 111000
        delta_lon_formatted = float("{:.3f}".format(delta_lon))
        kalman_delta_longitude.append(delta_lon_formatted)
        delta_lon = abs(fix_longitude[i] - particle_longitude[i]) * 111000
        delta_lon_formatted = float("{:.3f}".format(delta_lon))
        particle_delta_longitude.append(delta_lon_formatted)

        delta_lat = abs(fix_latitude[i] - fil_latitude[i]) * 111000
        delta_lat_formatted = float("{:.3f}".format(delta_lat))
        kalman_delta_latitude.append(delta_lat_formatted)
        delta_lat = abs(fix_latitude[i] - particle_latitude[i]) * 111000
        delta_lat_formatted = float("{:.3f}".format(delta_lat))
        particle_delta_latitude.append(delta_lat_formatted)

        kalman_delta_altitude.append(round(abs(fix_altitude[i] - fil_altitude[i]), 3))
        particle_delta_altitude.append(round(abs(fix_altitude[i] - particle_altitude[i]), 3))

        time_diff = round(time_epoch[i] - time_epoch[0], 3)
        delta_time.append(time_diff)
        std_kalman += math.sqrt(
            (fil_longitude[i] - fix_longitude[i]) ** 2 + (fil_latitude[i] - fix_latitude[i]) ** 2 + (
                    fil_altitude[i] - fix_altitude[i]) ** 2)
        std_particle += math.sqrt(
            (fil_longitude[i] - fix_longitude[i]) ** 2 + (particle_latitude[i] - particle_latitude[i]) ** 2 + (
                    particle_altitude[i] - particle_altitude[i]) ** 2)
        horizontal.append(
            math.sqrt((fil_longitude[i] - fix_longitude[i]) ** 2 + (fil_latitude[i] - fix_latitude[i]) ** 2))
        circle.append(
            math.sqrt((fil_longitude[i] - fix_longitude[i]) ** 2 + (fil_latitude[i] - fix_latitude[i]) ** 2 +
                      (fil_altitude[i] - fix_altitude[i]) ** 2))

    for i in range(len(vx_pixhawk)):
        delta_vx.append(abs(vx_kalman[i] - vx_pixhawk[i]))
        delta_vy.append(abs(vy_kalman[i] - vy_pixhawk[i]))
        delta_vz.append(abs(vz_kalman[i] - vz_pixhawk[i]))

    def quaternionToDegree(quaternion, title):
        roll_degrees, pitch_degrees, yaw_degrees = [], [], []
        if title != 'particle':
            for i in range(len(quaternion[0])):
                x_val = quaternion[0][i]
                y_val = quaternion[1][i]
                z_val = quaternion[2][i]
                w_val = quaternion[3][i]

                sinr_cosp = 2 * (w_val * x_val + y_val * z_val)
                cosr_cosp = 1 - 2 * (x_val * x_val + y_val * y_val)
                roll_rad = np.arctan2(sinr_cosp, cosr_cosp)

                sinp = 2 * (w_val * y_val - z_val * x_val)
                pitch_rad = np.arcsin(sinp)

                siny_cosp = 2 * (w_val * z_val + x_val * y_val)
                cosy_cosp = 1 - 2 * (y_val * y_val + z_val * z_val)
                yaw_rad = np.arctan2(siny_cosp, cosy_cosp)

                roll_deg = roll_rad * 180 / math.pi
                pitch_deg = pitch_rad * 180 / math.pi
                yaw_deg = yaw_rad * 180 / math.pi

                roll_degrees.append(roll_deg)
                pitch_degrees.append(pitch_deg)
                yaw_degrees.append(float(yaw_deg))
            return yaw_degrees, pitch_degrees, roll_degrees
        else:
            for i in range(len(quaternion)):
                yaw_rad = quaternion[i]
                yaw_deg = yaw_rad * 180 / math.pi
                yaw_degrees.append(yaw_deg)
        return yaw_degrees

    '''For pixhawk orientation angles'''
    pixhawk_yaw_degrees, pixhawk_pitch_degrees, pixhawk_roll_degrees = quaternionToDegree(pixhawk_quaternion, 'pixhawk')

    '''For Kalman orientation angles'''
    kalman_yaw_degrees, kalman_pitch_degrees, kalman_roll_degrees = quaternionToDegree(kalman_quaternion, 'kalman')

    # print(len(pixhawk_yaw_degrees), len(kalman_yaw_degrees))
    '''For Particle yaw angle'''
    particle_yaw_degrees = quaternionToDegree(particle_yaw, 'particle')
    particle_yaw_rms_degrees = quaternionToDegree(particle_yaw_rms, 'particle')

    # SD = math.sqrt(sd / len(fil_longitude))
    SD = 0

    return kalman_delta_longitude, kalman_delta_latitude, kalman_delta_altitude, particle_delta_longitude, particle_delta_latitude, particle_delta_altitude, \
        delta_time, horizontal, circle, SD, pixhawk_yaw_degrees, pixhawk_pitch_degrees, pixhawk_roll_degrees, kalman_yaw_degrees, kalman_pitch_degrees, kalman_roll_degrees, \
        particle_yaw_degrees, particle_yaw_rms_degrees, delta_vx, delta_vy, delta_vz


def writeToExcel(fix_longitude, fix_latitude, fix_altitude, time_epoch, h_acc, v_acc, fil_longitude, fil_latitude,
                 fil_altitude, particle_longitude, particle_latitude, particle_altitude, pixhawk_yaw_degrees,
                 pixhawk_pitch_degrees, pixhawk_roll_degrees,
                 particle_yaw, kalman_yaw_degrees, kalman_pitch_degrees, kalman_roll_degrees, kalman_delta_longitude,
                 kalman_delta_latitude, kalman_delta_altitude, particle_delta_longitude, particle_delta_latitude,
                 particle_delta_altitude, delta_time, horizontal, circle, SD, pressure, bag, output_directory):
    workbook = Workbook()
    sheet1 = workbook.active
    sheet1.title = 'gps_fix'
    sheet2 = workbook.create_sheet(title='gps_filtered (Kalman)')
    sheet3 = workbook.create_sheet(title='gps_filtered (Particle)')
    sheet4 = workbook.create_sheet(title='calculated_data')

    headers = ['longitude', 'latitude', 'altitude', 'time_epoch', 'h_acc', 'v_acc', 'yaw', 'pitch', 'roll', 'pressure',
               'min_long', 'min_lat', 'min_alt', 'max_long', 'max_lat', 'max_alt']

    # Writing data from arrays to the sheet 1 (gps fix sheet)
    for col, header in enumerate(headers, start=1):
        sheet1.cell(row=1, column=col, value=header)
    data = [fix_longitude, fix_latitude, fix_altitude, time_epoch, h_acc, v_acc, pixhawk_yaw_degrees,
            pixhawk_pitch_degrees, pixhawk_roll_degrees, pressure]
    for col, column_data in enumerate(data, start=1):
        for row, value in enumerate(column_data, start=2):
            sheet1.cell(row=row, column=col, value=value)

    # Writing single values to the sheet 1 (gps fix sheet)
    single_values = {'min_long': min(fix_longitude), 'min_lat': min(fix_latitude), 'min_alt': min(fix_altitude),
                     'max_long': max(fix_longitude), 'max_lat': max(fix_latitude), 'max_alt': max(fix_altitude)}
    for col, value in enumerate(single_values, start=len(data) + 1):
        sheet1.cell(row=1, column=col, value=value)
        sheet1.cell(row=2, column=col, value=single_values[value])

    headers = ['longitude', 'latitude', 'altitude', 'time_epoch', 'h_acc', 'v_acc', 'yaw', 'pitch', 'roll',
               'min_long', 'min_lat', 'min_alt', 'max_long', 'max_lat', 'max_alt']

    # Writing data from arrays to the sheet 2 (gps after EKF)
    for col, header in enumerate(headers, start=1):
        sheet2.cell(row=1, column=col, value=header)
    data_2 = [fil_longitude, fil_latitude, fil_altitude, time_epoch, h_acc, v_acc, kalman_yaw_degrees,
              kalman_pitch_degrees, kalman_roll_degrees]

    for col, column_data in enumerate(data_2, start=1):
        for row, value in enumerate(column_data, start=2):
            sheet2.cell(row=row, column=col, value=value)

    # Writing single values to the sheet 2 (gps after EKF)
    single_values_2 = {'min_long': min(fil_longitude), 'min_lat': min(fil_latitude), 'min_alt': min(fil_altitude),
                       'max_long': max(fil_longitude), 'max_lat': max(fil_latitude), 'max_alt': max(fil_altitude)}
    for col, value in enumerate(single_values_2, start=len(data) + 1):
        sheet2.cell(row=1, column=col, value=value)
        sheet2.cell(row=2, column=col, value=single_values_2[value])

    headers = ['longitude', 'latitude', 'altitude', 'time_epoch', 'h_acc', 'v_acc', 'yaw', 'min_long', 'min_lat',
               'min_alt', 'max_long', 'max_lat', 'max_alt']
    # Writing data from arrays to the sheet 3 (gps after Particle filter)

    for col, header in enumerate(headers, start=1):
        sheet3.cell(row=1, column=col, value=header)
    data_3 = [particle_longitude, particle_latitude, particle_altitude, time_epoch, h_acc, v_acc, particle_yaw]

    for col, column_data in enumerate(data_3, start=1):
        for row, value in enumerate(column_data, start=2):
            sheet3.cell(row=row, column=col, value=value)

    # Writing single values to the sheet 3 (gps after Particle filter)
    single_values_3 = {'min_long': min(particle_longitude), 'min_lat': min(particle_latitude),
                       'min_alt': min(particle_altitude),
                       'max_long': max(particle_longitude), 'max_lat': max(particle_latitude),
                       'max_alt': max(particle_altitude)}
    for col, value in enumerate(single_values_3, start=len(data) + 1):
        sheet3.cell(row=1, column=col, value=value)
        sheet3.cell(row=2, column=col, value=single_values_3[value])

    headers = ['delta_longitude', 'delta_latitude', 'delta_altitude', 'horizontal', 'circle']
    # Writing data from arrays to sheet 4 (calculated data)

    for col, header in enumerate(headers, start=1):
        sheet4.cell(row=1, column=col, value=header)
    data_4 = [kalman_delta_longitude, kalman_delta_latitude, kalman_delta_altitude, horizontal, circle]
    for col, column_data in enumerate(data_4, start=1):
        for row, value in enumerate(column_data, start=2):
            sheet4.cell(row=row, column=col, value=value)

    # Writing single values to sheet 4 (calculated data)
    single_values_4 = {'SD': SD, 'avg_delta_long, m': math.fsum(kalman_delta_longitude) / len(kalman_delta_longitude),
                       'avg_delta_lat, m':
                           math.fsum(kalman_delta_latitude) / len(kalman_delta_latitude),
                       'avg_delta_alt, m': math.fsum(kalman_delta_altitude) / len(kalman_delta_altitude)}
    for col, value in enumerate(single_values_4, start=len(data)):
        sheet4.cell(row=1, column=col, value=value)
        sheet4.cell(row=2, column=col, value=single_values_4[value])

    filepath_excel = f'{output_directory}/' + f'{bag_name_list[bag].split(".")[0]}' + '.xlsx'
    workbook.save(filepath_excel)


def latLong(latitude, longitude, title, output_directory):
    "Producing gps/fix or gps/filtered lat/long graph"
    plt.scatter(latitude, longitude, c='r', label='fix', s=5)
    plt.xlabel('Longitude')
    plt.ylabel('Latitude')
    plt.title('GPS' + f'{title}' + 'Lat Long graph')

    num_ticks_x = 6
    num_ticks_y = 6

    # Determine the range and step for the X and Y axes
    min_latitude = min(latitude)
    max_latitude = max(latitude)
    min_longitude = min(longitude)
    max_longitude = max(longitude)
    step_x = (max_longitude - min_longitude) / (num_ticks_x - 1) if (num_ticks_x - 1) != 0 else 1
    step_y = (max_latitude - min_latitude) / (num_ticks_y - 1) if (num_ticks_y - 1) != 0 else 1

    # Calculate the tick positions for both axes
    ticks_x = [min_longitude + i * step_x for i in range(num_ticks_x)]
    ticks_y = [min_latitude + i * step_y for i in range(num_ticks_y)]

    # Set values and labels on the X and Y axes
    plt.xticks(ticks_x)
    plt.gca().set_xticklabels([f'{tick:.7f}' for tick in ticks_x], fontsize=7)  # Уменьшаем размер шрифта

    plt.yticks(ticks_y)
    plt.gca().set_yticklabels([f'{tick:.7f}' for tick in ticks_y], fontsize=7)  # Уменьшаем размер шрифта

    # Add tick label manually if only one value is present on the X axis
    if len(longitude) == 1:
        plt.gca().set_xticklabels([f'{longitude[0]:.7f}'])

    # Add tick label manually if only one value is present on the Y axis
    if len(latitude) == 1:
        plt.gca().set_yticklabels([f'{latitude[0]:.7f}'])

    x_formatter = ticker.ScalarFormatter(useOffset=False)
    y_formatter = ticker.ScalarFormatter(useOffset=False)
    plt.gca().xaxis.set_major_formatter(x_formatter)
    plt.gca().yaxis.set_major_formatter(y_formatter)

    image = plt.gca().get_title() + '.png'
    filepath = os.path.join(output_directory, image)
    plt.tight_layout()
    plt.savefig(filepath, dpi=200)
    plt.close()
    return image


def sameCoordinatePlaneLatLongGraph(fix_longitude, fix_latitude, type_of_filter_longitude, type_of_filter_latitude,
                                    title, output_directory):
    '''Producing gps/filtered, gps/fix lat/long graph on the same coordinate plane'''
    fig, ax = plt.subplots()

    # Make 1st graph
    plt.scatter(fix_longitude, fix_latitude, c='r', label='fix', s=5)

    # Make 2nd graph
    plt.scatter(type_of_filter_longitude, type_of_filter_latitude, c='b', label=f'{title}', s=5)

    plt.xlabel('Longitude')
    plt.ylabel('Latitude')
    plt.title('GPS fix and GPS ' + f'{title}' + ' Lat Long graph')

    num_ticks_x = 6
    num_ticks_y = 6

    # Determine the range and step for the X and Y axes of both graphs
    min_longitude = min(min(fix_longitude), min(type_of_filter_longitude))
    max_longitude = max(max(fix_longitude), max(type_of_filter_longitude))
    step_x = (max_longitude - min_longitude) / (num_ticks_x - 1)

    min_latitude = min(min(fix_latitude), min(type_of_filter_latitude))
    max_latitude = max(max(fix_latitude), max(type_of_filter_latitude))
    step_y = (max_latitude - min_latitude) / (num_ticks_y - 1)

    # Calculate the tick positions for both graphs
    ticks_x = [min_longitude + i * step_x for i in range(num_ticks_x)]
    ticks_y = [min_latitude + i * step_y for i in range(num_ticks_y)]

    # Set values and labels on the X and Y axes for both graphs
    plt.xticks(ticks_x)
    plt.gca().set_xticklabels([f'{tick:.7f}' for tick in ticks_x], fontsize=8)  # Уменьшаем размер шрифта

    plt.yticks(ticks_y)
    plt.gca().set_yticklabels([f'{tick:.7f}' for tick in ticks_y], fontsize=8)  # Уменьшаем размер шрифта

    x_formatter = ticker.ScalarFormatter(useOffset=False)
    y_formatter = ticker.ScalarFormatter(useOffset=False)
    plt.gca().xaxis.set_major_formatter(x_formatter)
    plt.gca().yaxis.set_major_formatter(y_formatter)

    # Legend configure
    plt.legend()
    plt.legend(loc='best')
    plt.tight_layout()

    image = plt.gca().get_title() + '.png'
    filepath = os.path.join(output_directory, image)
    # Show graph
    plt.savefig(filepath, dpi=200)
    plt.close()
    return image


def combineGPSGraphs(SD, fix_longitude, fix_latitude, fil_longitude, fil_latitude, output_directory):
    # Uploading all images to crop them into 1 image
    image1 = Image.open(output_directory + '/' + latLong(fix_longitude, fix_latitude, ' fix ', output_directory))
    # image2 = Image.open(output_directory + '/' + latLong(fil_longitude, fil_latitude, ' Kalman ', output_directory))
    image3 = Image.open(output_directory + '/' + sameCoordinatePlaneLatLongGraph(fix_longitude, fix_latitude,
                                                                                 fil_longitude, fil_latitude, 'Kalman',
                                                                                 output_directory))
    # image4 = Image.open(output_directory + '/' + latLong(particle_longitude, particle_latitude, ' Particle ', output_directory))
    image5 = Image.open(output_directory + '/' + sameCoordinatePlaneLatLongGraph(fix_longitude, fix_latitude,
                                                                                 particle_longitude, particle_latitude,
                                                                                 'Particle',
                                                                                 output_directory))

    # Define the size of the final image
    width = 2 * image1.width
    height = 2 * image1.height
    result_image_kalman = Image.new('RGB', (width, height), color="white")
    result_image_particle = Image.new('RGB', (width, height), color="white")

    # Inserting images to the final image Kalman filtered /fix
    result_image_kalman.paste(image1, (0, 0))
    result_image_kalman.paste(image5, (image1.width, 0))
    result_image_kalman.paste(image3, (0, image1.height))

    # result_image_particle.paste(image1, (0, 0))
    # result_image_particle.paste(image4, (image1.width, 0))
    # result_image_particle.paste(image5, (0, image1.height))

    # Add SD value to the final image
    # draw = ImageDraw.Draw(result_image)
    # font_size = 72
    # SD = "{:.4f}".format(SD)
    # font = ImageFont.truetype('arial.ttf', size=font_size)
    # text = f'SD: {SD}'
    # text_color = (255, 255, 255)
    # draw.text((4000, 4500), text, fill=(0, 0, 0), font=font)

    # Save final result
    result_image_kalman.save(output_directory + '/' + 'combined_gps.png')
    # result_image_particle.save(output_directory + '/' + 'fix, Kalman, Particle filtered.png')
    return result_image_kalman


def deltaPlot(coordinate, delta_time, title, output_directory):
    plt.scatter(delta_time, coordinate, c='b', label='fix', s=5)
    plt.xlabel('Time, s')
    plt.ylabel('Error, m')
    plt.title(f'{title}, m ')

    # Determine the number of values on the Y-axis
    num_ticks = 6
    step = (max(coordinate) - min(coordinate)) / (num_ticks - 1)
    if step == 0:
        step = 0.001
    ticks = np.arange(min(coordinate), max(coordinate) + step, step)

    # Set values and labels on the Y axis
    plt.yticks(ticks)
    plt.gca().set_yticklabels([f'{tick:.2f}' for tick in ticks])

    # Set the indentation for labels on the Y axis
    plt.gca().yaxis.set_tick_params(pad=20)

    image = plt.gca().get_title() + '.png'
    filepath = os.path.join(output_directory, image)
    plt.tight_layout()
    plt.savefig(filepath, dpi=200)
    plt.close()
    return image


def horizon(coordinate, delta_time, output_directory):
    fig, ax = plt.subplots()
    ax.scatter(delta_time, coordinate, c='b', label='fix', s=5)
    ax.set_xlabel('Time, s')
    ax.set_ylabel('Horizon')
    title = 'Horizontal'
    ax.set_title(f'{title}')
    # Number of values on Y axis
    num_ticks = 10
    # Calculate the step between values on the y-axis
    step = (max(coordinate) - min(coordinate)) / (num_ticks - 1)
    # Create an array with a given number of values on the y-axis
    ticks = np.linspace(min(coordinate), max(coordinate), num_ticks)
    ax.set_yticks(ticks)
    # Set the labels on the Y axis
    ax.set_yticklabels([f'{tick:.2f}' for tick in ticks])
    # Disable the use of offset
    ax.yaxis.set_major_formatter(plt.ScalarFormatter(useOffset=False))
    # Calculation and setting of optimal indents for chart elements
    plt.tight_layout()
    image = ax.get_title() + '.png'
    filepath = os.path.join(output_directory, image)
    plt.savefig(filepath, dpi=200)
    plt.close()
    return image


def CombindeDeltaGraphs(kalman_delta_latitude, kalman_delta_longitude, kalman_delta_altitude, particle_delta_latitude,
                        particle_delta_longitude,
                        particle_delta_altitude, delta_time, horizontal, output_directory):
    image1 = Image.open(
        output_directory + '/' + deltaPlot(kalman_delta_latitude, delta_time, "Fix, Kalman, difference in latitude",
                                           output_directory))
    image2 = Image.open(
        output_directory + '/' + deltaPlot(kalman_delta_longitude, delta_time, "Fix, Kalman, difference in longitude",
                                           output_directory))
    image3 = Image.open(
        output_directory + '/' + deltaPlot(kalman_delta_altitude, delta_time, "Fix, Kalman, difference in altitude",
                                           output_directory))

    image4 = Image.open(
        output_directory + '/' + deltaPlot(particle_delta_latitude, delta_time, "Fix, Particle, difference in latitude",
                                           output_directory))
    image5 = Image.open(output_directory + '/' + deltaPlot(particle_delta_longitude, delta_time,
                                                           "Fix, Particle, difference in longitude", output_directory))
    image6 = Image.open(
        output_directory + '/' + deltaPlot(particle_delta_altitude, delta_time, "Fix, Particle, difference in altitude",
                                           output_directory))

    # image4 = Image.open(output_directory + '/' + horizon(horizontal, delta_time, output_directory))
    width = 2 * image1.width
    height = 2 * image1.height
    result_image_ll = Image.new('RGB', (width, height), color="white")
    # Inserting images in the final image
    result_image_ll.paste(image1, (0, 0))
    result_image_ll.paste(image4, (image1.width, 0))
    result_image_ll.paste(image2, (0, image1.height))
    result_image_ll.paste(image5, (image1.width, image1.height))

    result_image_alt = Image.new('RGB', (width, height), color="white")
    # Inserting images in the final image
    result_image_alt.paste(image3, (0, 0))
    result_image_alt.paste(image6, (image1.width, 0))
    result_image_ll.save(output_directory + '/' + 'Kalman, Particle ll.png')
    result_image_alt.save(output_directory + '/' + 'Kalman, Particle alt.png')
    return result_image_ll, result_image_alt


def ThreeDPlot(long, lat, alt):
    fig = plt.figure()
    title = '3d delta plot'
    plt.title(f'{title}')
    ax = fig.add_subplot(111, projection='3d')
    ax.scatter(long, lat, alt)
    x_formatter = ticker.ScalarFormatter(useOffset=False)
    y_formatter = ticker.ScalarFormatter(useOffset=False)
    ax.xaxis.set_major_formatter(x_formatter)
    ax.yaxis.set_major_formatter(y_formatter)
    ax.set_xlabel('Longitude')
    ax.set_ylabel('Latitude')
    ax.set_zlabel('Altitude')
    image = plt.gca().get_title() + '.png'
    filepath = os.path.join(output_directory, image)
    plt.tight_layout()
    plt.savefig(filepath, dpi=200)


def circle_graph(coordinate, delta_time, output_directory):
    plt.scatter(delta_time, coordinate, c='b', label='fix', s=5)
    plt.xlabel('Time, s')
    plt.ylabel('Circle')
    title = 'Circle'
    plt.title(f'{title}')
    x_formatter = ticker.ScalarFormatter(useOffset=False)
    y_formatter = ticker.ScalarFormatter(useOffset=False)
    plt.gca().xaxis.set_major_formatter(x_formatter)
    plt.gca().yaxis.set_major_formatter(y_formatter)
    image = plt.gca().get_title() + '.png'
    filepath = os.path.join(output_directory, image)
    plt.tight_layout()
    plt.savefig(filepath, dpi=200)
    # plt.show()
    plt.close()


def accuracy_plot(coordinate, delta_time, title, output_directory):
    plt.scatter(delta_time, coordinate, c='b', label='fix', s=5)
    plt.xlabel('Time, s')
    plt.ylabel(f'{title}, mm')
    plt.title('GPS ' + f'{title}')
    # print(len(delta_time), len(h_acc))
    if len(delta_time) < len(h_acc):
        del h_acc[len(h_acc) - 1]
        del v_acc[len(v_acc) - 1]
    # x_formatter = ticker.ScalarFormatter(useOffset=False)
    # y_formatter = ticker.ScalarFormatter(useOffset=False)
    # plt.gca().xaxis.set_major_formatter(x_formatter)
    # plt.gca().yaxis.set_major_formatter(y_formatter)
    image = plt.gca().get_title() + '.png'
    filepath = os.path.join(output_directory, image)
    # plt.tight_layout()
    plt.locator_params(axis='y', nbins=20, tight=True, steps=[0.5])
    plt.savefig(filepath, dpi=200)
    # plt.locator_params(axis='y', tight=True, nbins=3)
    plt.yticks([0, 1, 2, 3, 4, 5])
    plt.gca().yaxis.set_tick_params(pad=5)
    plt.show()
    # plt.close()
    return image


def combine_accuracy_graphs(h_acc, v_acc, output_directory):
    print(len(h_acc))
    image1 = Image.open(
        output_directory + '/' + accuracy_plot(h_acc, delta_time, "horizontal accuracy", output_directory))
    image2 = Image.open(
        output_directory + '/' + accuracy_plot(v_acc, delta_time, "vertical accuracy", output_directory))
    width = 2 * image1.width
    height = 2 * image1.height
    result_image = Image.new('RGB', (width, height), color="white")
    # Inserting images in the final image
    result_image.paste(image1, (0, 0))
    result_image.paste(image2, (image1.width, 0))
    result_image.save(output_directory + 'combined_accuracy.png')


def plot_angle(sourceOne, sourceTwo, sourceThree, sourceNameOne, sourceNameTwo, sourceNameThree, angle, sourceFour=None,
               sourceNameFour=None):
    if angle == 'yaw':
        plt.scatter(np.arange(len(sourceOne)), sourceOne, c='r', label=sourceNameOne, s=5)
        plt.scatter(np.arange(len(sourceTwo)), sourceTwo, c='b', label=sourceNameTwo, s=5)
        plt.scatter(np.arange(len(sourceThree)), sourceThree, c='g', label=sourceNameThree, s=5)
        plt.scatter(np.arange(len(sourceFour)), sourceFour, c='black', label=sourceNameFour, s=5)

        plt.xlabel('Index')
        plt.ylabel(f'{angle}, degress')
        plt.title(
            f'{sourceNameOne}' + ", " + f'{sourceNameTwo}' + ", " + f'{sourceNameThree}' + ", " + f'{sourceNameFour}' + ", " + f'{angle}')
        plt.legend()
        plt.legend(loc='best')
        plt.tight_layout()
    else:
        plt.scatter(np.arange(len(sourceOne)), sourceOne, c='r', label=sourceNameOne, s=5)
        plt.scatter(np.arange(len(sourceTwo)), sourceTwo, c='b', label=sourceNameTwo, s=5)
        plt.scatter(np.arange(len(sourceThree)), sourceThree, c='g', label=sourceNameThree, s=5)
        plt.xlabel('Index')
        plt.ylabel(f'{angle}, degress')
        plt.title(f'{sourceNameOne}' + ", " + f'{sourceNameTwo}' + ", " + f'{sourceNameThree}' + ", " + f'{angle}')
        plt.legend()
        plt.legend(loc='best')
        plt.tight_layout()

    image = plt.gca().get_title() + '.png'
    filepath = os.path.join(output_directory, image)
    plt.savefig(filepath, dpi=200)
    # plt.show()
    plt.close()
    return image


def combineAnglesGraphsDegrees():
    '''Uploading all images to crop them into 1 image
    XYZ and YPR - X - pitch, Y - roll, Z - yaw
    '''
    image1 = Image.open(
        output_directory + '/' + plot_angle(pixhawk_yaw_degrees, kalman_yaw_degrees, zmag, 'pixhawk', 'kalman',
                                            'magnetometer', 'yaw', particle_yaw, 'particle'))
    image2 = Image.open(
        output_directory + '/' + plot_angle(pixhawk_pitch_degrees, kalman_pitch_degrees, xmag, 'pixhawk', 'kalman',
                                            'magnetometer', 'pitch'))
    image3 = Image.open(
        output_directory + '/' + plot_angle(pixhawk_roll_degrees, kalman_roll_degrees, ymag, 'pixhawk', 'kalman',
                                            'magnetometer', 'roll'))
    width = 2 * image1.width
    height = 2 * image1.height
    result_image_angles = Image.new('RGB', (width, height), color="white")
    result_image_angles.paste(image1, (0, 0))
    result_image_angles.paste(image2, (image1.width, 0))
    result_image_angles.paste(image3, (0, image1.height))
    result_image_angles.save(output_directory + '/' + 'Kalman, Particle, Pixhawk, magnetometer orientation.png')
    return result_image_angles


def speedPlot(sourceOne, sourceNameOne, sourceTwo, sourceNameTwo, sourceThree, sourceNameThree, title):
    plt.scatter(np.arange(len(sourceOne)), sourceOne, c='r', label=sourceNameOne, s=7)
    plt.scatter(np.arange(len(sourceTwo)), sourceTwo, c='b', label=sourceNameTwo, s=2)
    plt.scatter(np.arange(len(sourceThree)), sourceThree, c='g', label=sourceNameThree, s=5)
    plt.xlabel('Index')
    plt.ylabel(f'{title}')
    plt.title(f'{title}')
    num_ticks = 12
    '''Expiremental feature - combine all arrays to one array to have more representable Y-axis labels'''
    commonArray = np.concatenate((sourceOne, sourceTwo, sourceThree))
    step = (max(commonArray) - min(commonArray)) / (num_ticks - 1)
    ticks = np.arange(min(commonArray), max(commonArray) + step, step)

    # Set values and labels on the Y axis
    plt.yticks(ticks)
    plt.gca().set_yticklabels([f'{tick:.2f}' for tick in ticks], fontsize=8)
    plt.legend(loc='best')
    plt.tight_layout()

    image = plt.gca().get_title().replace('/', '_') + '.png'
    filepath = os.path.join(output_directory, image)
    plt.savefig(filepath, dpi=200)
    plt.close()
    return image


def rmsPlot(sourceOne, sourceNameOne, title, sourceTwo=None, sourceNameTwo=None, sourceThree=None,
            sourceNameThree=None):
    # sourceOneMetres, sourceTwoMetres = [], []
    num_ticks = 6
    if sourceNameOne != 'yaw':
        plt.scatter(np.arange(len(sourceOne)), sourceOne, c='r', label=sourceNameOne, s=5)
        plt.scatter(np.arange(len(sourceTwo)), sourceTwo, c='b', label=sourceNameTwo, s=5)
        plt.scatter(np.arange(len(sourceThree)), sourceThree, c='g', label=sourceNameThree, s=5)
        plt.xlabel('Index')
        plt.ylabel(f'{title}')
        plt.title(f'{title}')
        commonArray = np.concatenate((sourceOne, sourceTwo, sourceThree))
        step = 2 * (max(commonArray) - min(commonArray)) / (num_ticks - 1)
        ticks = np.arange(min(commonArray), max(commonArray) + step, step)

    else:
        plt.scatter(np.arange(len(sourceOne)), sourceOne, c='r', label=sourceNameOne, s=5)
        plt.xlabel('Index')
        plt.ylabel('yaw RMS')
        plt.title(f'{title}' + ' ' + f'{sourceNameOne}' + ", degress")
        step = 2 * (max(sourceOne) - min(sourceOne)) / (num_ticks - 1)
        ticks = np.arange(min(sourceOne), max(sourceOne) + step, step)

    # Set values and labels on the Y axis
    plt.yticks(ticks)
    plt.gca().set_yticklabels([f'{tick:.3f}' for tick in ticks], fontsize=8)
    plt.legend(loc='best')
    plt.tight_layout()
    image = plt.gca().get_title() + '.png'
    filepath = os.path.join(output_directory, image)
    plt.savefig(filepath, dpi=200)
    plt.close()
    return image


def speedDeltaPlot(sourceOne, sourceNameOne, sourceTwo, sourceNameTwo, sourceThree, sourceNameThree, title):
    plt.scatter(np.arange(len(sourceOne)), sourceOne, c='r', label=sourceNameOne, s=7)
    plt.scatter(np.arange(len(sourceTwo)), sourceTwo, c='b', label=sourceNameTwo, s=2)
    plt.scatter(np.arange(len(sourceThree)), sourceThree, c='g', label=sourceNameThree, s=5)
    plt.xlabel('Index')
    plt.ylabel(f'{title}')
    plt.title(f'{title}')
    num_ticks = 12
    '''Expiremental feature - combine all arrays to one array to have more representable Y-axis labels'''
    commonArray = np.concatenate((sourceOne, sourceTwo, sourceThree))
    step = (max(commonArray) - min(commonArray)) / (num_ticks - 1)
    ticks = np.arange(min(commonArray), max(commonArray) + step, step)

    # Set values and labels on the Y axis
    plt.yticks(ticks)
    plt.gca().set_yticklabels([f'{tick:.2f}' for tick in ticks], fontsize=8)
    plt.legend(loc='best')
    plt.tight_layout()

    image = plt.gca().get_title().replace('/', '_') + '.png'
    filepath = os.path.join(output_directory, image)
    plt.savefig(filepath, dpi=200)
    plt.close()
    return image


def pdfOutput():  # vx_kalman_rms, vy_kalman_rms, vz_kalman_rms
    anglesImage = combineAnglesGraphsDegrees()
    coordintatesImage = combineGPSGraphs(SD, fix_longitude, fix_latitude, fil_longitude, fil_latitude, output_directory)
    deltaLlImage, deltaAltImage = CombindeDeltaGraphs(kalman_delta_latitude, kalman_delta_longitude,
                                                      kalman_delta_altitude, particle_delta_latitude,
                                                      particle_delta_longitude,
                                                      particle_delta_altitude, delta_time, horizontal, output_directory)
    linearVelocitiesKalman = Image.open(
        output_directory + '/' + speedPlot(vx_kalman, 'vx', vy_kalman, 'vy', vz_kalman, 'vz',
                                           'Kalman linear speed, m/s'))
    angularVeloctiesKalman = Image.open(
        output_directory + '/' + speedPlot(wx_kalman, 'wx', wy_kalman, 'wy', wz_kalman, 'wz',
                                           'Kalman angular speed, rad/s'))
    linearVelocitiesPixhawk = Image.open(
        output_directory + '/' + speedPlot(vx_pixhawk, 'vx', vy_pixhawk, 'vy', vz_pixhawk, 'vz',
                                           'Pixhawk linear speed, m/s'))

    rmsParticleLLAImage = Image.open(
        output_directory + '/' + rmsPlot(particle_lat_rms, 'lat rms', 'Particle LLA RMS', particle_long_rms, 'long rms',
                                         particle_alt_rms, 'alt rms'))
    rmsParticleYawImage = Image.open(output_directory + '/' + rmsPlot(particle_yaw_rms_degrees, 'yaw', 'particle'))
    rmsKalmanLLAImage = Image.open(
        output_directory + '/' + rmsPlot(kalman_lat_rms, 'lat rms', 'Kalman LLA RMS', kalman_long_rms, 'long rms',
                                         kalman_alt_rms, 'alt rms'))
    rmsKalmanLinImage = Image.open(
        output_directory + '/' + rmsPlot(vx_kalman_rms, 'vx rms', 'Kalman LV RMS', vy_kalman_rms, 'vy rms',
                                         vz_kalman_rms, 'vz rms'))
    rmsKalmanAngImage = Image.open(
        output_directory + '/' + rmsPlot(wx_kalman_rms, 'wx rms', 'Kalman AV RMS', wy_kalman_rms, 'wy rms',
                                         wz_kalman_rms, 'wz rms'))
    deltaVelImage = Image.open(
        output_directory + '/' + speedDeltaPlot(delta_vx, 'delta vx', delta_vy, 'delta vy', delta_vz, 'delta vz',
                                                'Delta in linear velocities in m/s between Kalman and Pixhawk'))

    width = 2 * rmsParticleLLAImage.width
    height = 2 * rmsParticleLLAImage.height

    result_image_rms = Image.new('RGB', (width, height), color="white")
    result_image_rms.paste(rmsParticleLLAImage, (0, 0))
    result_image_rms.paste(rmsParticleYawImage, (rmsParticleLLAImage.width, 0))
    result_image_rms.paste(rmsKalmanLLAImage, (0, rmsParticleYawImage.height))

    result_image_velocities = Image.new('RGB', (width, height), color="white")
    result_image_velocities.paste(linearVelocitiesKalman, (0, 0))
    result_image_velocities.paste(angularVeloctiesKalman, (linearVelocitiesKalman.width, 0))
    result_image_velocities.paste(linearVelocitiesPixhawk, (0, linearVelocitiesKalman.height))

    result_image_velocities_rms = Image.new('RGB', (width, height), color="white")
    result_image_velocities_rms.paste(rmsKalmanLinImage, (0, 0))
    result_image_velocities_rms.paste(rmsKalmanAngImage, (rmsKalmanLinImage.width, 0))
    result_image_velocities_rms.paste(deltaVelImage, (0, rmsKalmanLinImage.height))

    im_1 = coordintatesImage.convert('RGB')
    im_2 = deltaLlImage.convert('RGB')
    im_3 = deltaAltImage.convert('RGB')
    im_4 = anglesImage.convert('RGB')
    im_5 = result_image_rms.convert('RGB')
    im_6 = result_image_velocities.convert('RGB')
    im_7 = result_image_velocities_rms.convert('RGB')

    image_list = [im_2, im_3, im_4, im_5, im_6, im_7]
    im_1.save(output_directory + "/Report.pdf", save_all=True, append_images=image_list)


if __name__ == '__main__':
    bag_name_list = ['rosbag2_2023_08_13-11_42_26']  # rosbag2_2023_08_13-11_31_55, rosbag2_2023_08_13-11_37_50
    for bag in tqdm(range(len(bag_name_list))):
        input_directory = './input/' + bag_name_list[bag]
        output_directory = './output/' + (bag_name_list[bag]).split('.')[0]
        os.makedirs(output_directory, exist_ok=True)

        fix_longitude, fix_latitude, fix_altitude, time_epoch, fil_longitude, fil_latitude, fil_altitude, h_acc, \
            v_acc, xmag, ymag, zmag, w_pixhawk, x_pixhawk, y_pixhawk, z_pixhawk, w_kalman, x_kalman, y_kalman, z_kalman, angular_velocity_x, angular_velocity_y, \
            angular_velocity_z, linear_acceleration_x, linear_acceleration_y, linear_acceleration_z, pressure, particle_longitude, particle_latitude, \
            particle_altitude, particle_yaw, particle_lat_rms, particle_long_rms, particle_alt_rms, particle_yaw_rms, kalman_lat_rms, kalman_long_rms, \
            kalman_alt_rms, vx_kalman, vy_kalman, vz_kalman, wx_kalman, wy_kalman, wz_kalman, vx_kalman_rms, vy_kalman_rms, vz_kalman_rms, wx_kalman_rms, \
            wy_kalman_rms, wz_kalman_rms, vx_pixhawk, vy_pixhawk, vz_pixhawk = \
            extractData(output_directory)

        kalman_delta_longitude, kalman_delta_latitude, kalman_delta_altitude, particle_delta_longitude, particle_delta_latitude, particle_delta_altitude, \
            delta_time, horizontal, circle, SD, pixhawk_yaw_degrees, pixhawk_pitch_degrees, pixhawk_roll_degrees, kalman_yaw_degrees, kalman_pitch_degrees, kalman_roll_degrees, \
            particle_yaw_degrees, particle_yaw_rms_degrees, delta_vx, delta_vy, delta_vz = calculations(fix_longitude,
                                                                                                        fix_latitude,
                                                                                                        fix_altitude,
                                                                                                        time_epoch,
                                                                                                        fil_longitude,
                                                                                                        fil_latitude,
                                                                                                        fil_altitude,
                                                                                                        x_pixhawk,
                                                                                                        y_pixhawk,
                                                                                                        z_pixhawk,
                                                                                                        w_pixhawk,
                                                                                                        x_kalman,
                                                                                                        y_kalman,
                                                                                                        z_kalman,
                                                                                                        w_kalman,
                                                                                                        particle_yaw,
                                                                                                        particle_yaw_rms)

        writeToExcel(fix_longitude, fix_latitude, fix_altitude, time_epoch, h_acc, v_acc, fil_longitude, fil_latitude,
                     fil_altitude, particle_longitude, particle_latitude, particle_altitude,
                     pixhawk_yaw_degrees, pixhawk_pitch_degrees, pixhawk_roll_degrees, particle_yaw, kalman_yaw_degrees,
                     kalman_pitch_degrees,
                     kalman_roll_degrees, kalman_delta_longitude, kalman_delta_latitude, kalman_delta_altitude,
                     particle_delta_longitude,
                     particle_delta_latitude, particle_delta_altitude, delta_time, horizontal, circle, SD, pressure,
                     bag, output_directory)

        pdfOutput()

